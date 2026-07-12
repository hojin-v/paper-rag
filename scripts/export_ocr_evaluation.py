from __future__ import annotations

import argparse
import json
import re
from collections import Counter
from collections.abc import Iterable, Sequence
from pathlib import Path
from typing import Any

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.worksheet import Worksheet
from sqlalchemy import bindparam, text
from sqlalchemy.engine import Engine, RowMapping

from paperrag.config import Settings, get_settings
from paperrag.db import get_engine
from paperrag.review.models import ReviewDocument
from paperrag.review.store import FileReviewStore

EXCEL_CELL_LIMIT = 32_767
HEADER_FILL = PatternFill("solid", fgColor="1F4E78")
CORRECTED_FILL = PatternFill("solid", fgColor="FFF2CC")
WARNING_FILL = PatternFill("solid", fgColor="F4CCCC")
HAN_RE = re.compile(r"[\u4e00-\u9fff]")
BENGALI_RE = re.compile(r"[\u0980-\u09ff]")
FORMAT_POLLUTION_RE = re.compile(r"```|抱歉|JSON回答|已结束|以下是结果", re.IGNORECASE)


def parse_args() -> argparse.Namespace:
    parser = argparse.ArgumentParser(
        description="검수 문서의 레이아웃·OCR·LLM 결과를 평가용 엑셀로 내보냅니다."
    )
    parser.add_argument(
        "--document-id",
        action="append",
        required=True,
        dest="document_ids",
        help="내보낼 검수 문서 ID. 여러 번 지정할 수 있습니다.",
    )
    parser.add_argument("--output", type=Path, help="생성할 .xlsx 파일 경로")
    return parser.parse_args()


def main() -> int:
    args = parse_args()
    settings = get_settings()
    output = args.output or settings.result_dir / "evaluations" / "ocr-evaluation.xlsx"
    documents = _load_ingested_documents(settings, args.document_ids)
    paper_ids = [int(document.paper_id) for document in documents if document.paper_id]
    data = _load_database_rows(get_engine(settings), paper_ids)
    output.parent.mkdir(parents=True, exist_ok=True)
    _build_workbook(settings, documents, data).save(output)
    report_path = output.with_suffix(".md")
    report_path.write_text(
        _build_markdown(settings, documents, data, output), encoding="utf-8"
    )
    print(output.resolve())
    print(report_path.resolve())
    return 0


def _load_ingested_documents(
    settings: Settings, document_ids: Sequence[str]
) -> list[ReviewDocument]:
    store = FileReviewStore(settings.review_dir)
    documents = [store.get(document_id) for document_id in document_ids]
    incomplete = [document.document_id for document in documents if document.paper_id is None]
    if incomplete:
        raise RuntimeError(f"아직 적재되지 않은 문서입니다: {', '.join(incomplete)}")
    return documents


def _load_database_rows(engine: Engine, paper_ids: Sequence[int]) -> dict[str, list[RowMapping]]:
    if not paper_ids:
        return {"papers": [], "paragraphs": [], "keywords": [], "tables": []}
    statements = {
        "papers": """
            SELECT paper_id, title, authors, published_year, journal, abstract,
                   abstract_summary, source_file_path, status
            FROM papers WHERE paper_id IN :paper_ids ORDER BY paper_id
        """,
        "paragraphs": """
            SELECT paragraph_id, paper_id, section_name, paragraph_order,
                   original_text, cleaned_text, summary, keywords, is_topic_relevant
            FROM paragraphs WHERE paper_id IN :paper_ids
            ORDER BY paper_id, paragraph_order
        """,
        "keywords": """
            SELECT pk.paper_id, k.keyword, k.display_form, k.frequency, pk.score
            FROM paper_keywords pk JOIN keywords k ON k.keyword_id = pk.keyword_id
            WHERE pk.paper_id IN :paper_ids
            ORDER BY pk.paper_id, pk.score DESC, k.display_form
        """,
        "tables": """
            SELECT table_id, paper_id, table_title, table_text, table_summary
            FROM paper_tables WHERE paper_id IN :paper_ids
            ORDER BY paper_id, table_id
        """,
    }
    rows: dict[str, list[RowMapping]] = {}
    with engine.connect() as connection:
        for name, sql in statements.items():
            statement = text(sql).bindparams(bindparam("paper_ids", expanding=True))
            rows[name] = list(
                connection.execute(statement, {"paper_ids": list(paper_ids)}).mappings()
            )
    return rows


def _build_workbook(
    settings: Settings,
    documents: Sequence[ReviewDocument],
    data: dict[str, list[RowMapping]],
) -> Workbook:
    workbook = Workbook()
    overview = workbook.active
    overview.title = "논문요약"
    paper_by_id = {int(row["paper_id"]): row for row in data["papers"]}
    paragraphs_by_paper = _group_by_paper(data["paragraphs"])
    keywords_by_paper = _group_by_paper(data["keywords"])
    tables_by_paper = _group_by_paper(data["tables"])

    overview_headers = [
        "문서 ID", "논문 ID", "파일명", "DB 제목", "저자", "페이지", "레이아웃 블록",
        "영역 유형별 개수", "검수 교정 블록", "평균 신뢰도", "초록 원문", "초록 요약",
        "대표 키워드", "본문 단락", "관련 단락", "요약 경고 단락", "표", "브라우저 검수 URL", "자동 점검 메모",
    ]
    overview.append(overview_headers)
    for document in documents:
        paper_id = int(document.paper_id or 0)
        paper = paper_by_id[paper_id]
        paragraphs = paragraphs_by_paper.get(paper_id, [])
        keywords = keywords_by_paper.get(paper_id, [])
        tables = tables_by_paper.get(paper_id, [])
        confidences = [
            float(block.confidence)
            for block in document.blocks
            if block.confidence is not None
        ]
        counts = Counter(block.detected_block_type or block.block_type for block in document.blocks)
        corrected = sum(block.review_status == "corrected" for block in document.blocks)
        overview.append(
            [
                document.document_id,
                paper_id,
                document.filename,
                paper["title"],
                paper["authors"],
                len(document.pages),
                len(document.blocks),
                json.dumps(dict(sorted(counts.items())), ensure_ascii=False),
                corrected,
                round(sum(confidences) / len(confidences), 4) if confidences else None,
                _excel_text(paper["abstract"]),
                _excel_text(paper["abstract_summary"]),
                ", ".join(str(row["display_form"]) for row in keywords),
                len(paragraphs),
                sum(bool(row["is_topic_relevant"]) for row in paragraphs),
                sum(bool(_summary_warning(str(row["summary"] or ""))) for row in paragraphs),
                len(tables),
                f"{settings.public_api_base_url.rstrip('/')}/documents/{document.document_id}/viewer",
                _quality_note(document, paper),
            ]
        )
    _style_sheet(overview, freeze="A2", wrap=True, widths={3: 34, 4: 48, 11: 55, 12: 55, 18: 70, 19: 58})

    layout = workbook.create_sheet("레이아웃_OCR")
    layout.append(
        [
            "문서 ID", "파일명", "페이지", "순서", "블록 ID", "자동 인식 유형", "현재 유형",
            "검수 상태", "자동 bbox", "현재 bbox", "신뢰도", "OCR 엔진", "원시 OCR", "교정 OCR", "적재 사용 텍스트",
        ]
    )
    for document in documents:
        for block in sorted(document.blocks, key=lambda item: item.order):
            layout.append(
                [
                    document.document_id,
                    document.filename,
                    block.page,
                    block.order,
                    block.block_id,
                    block.detected_block_type or block.block_type,
                    block.block_type,
                    block.review_status,
                    _bbox_text(block.detected_bbox or block.bbox),
                    _bbox_text(block.bbox),
                    block.confidence,
                    block.ocr_engine,
                    _excel_text(block.ocr_text),
                    _excel_text(block.corrected_text),
                    _excel_text(block.effective_text),
                ]
            )
            if block.review_status == "corrected":
                for cell in layout[layout.max_row]:
                    cell.fill = CORRECTED_FILL
    _style_sheet(layout, freeze="A2", wrap=True, widths={2: 38, 13: 70, 14: 70, 15: 70})

    paragraphs_sheet = workbook.create_sheet("단락_요약")
    paragraphs_sheet.append(
        [
            "논문 ID", "제목", "단락 ID", "섹션", "단락 순서", "원문", "정제 본문",
            "LLM 요약", "단락 키워드", "본문 관련 여부", "요약 자동 경고",
        ]
    )
    for row in data["paragraphs"]:
        paper = paper_by_id[int(row["paper_id"])]
        warning = _summary_warning(str(row["summary"] or ""))
        paragraphs_sheet.append(
            [
                row["paper_id"], paper["title"], row["paragraph_id"], row["section_name"],
                row["paragraph_order"], _excel_text(row["original_text"]),
                _excel_text(row["cleaned_text"]), _excel_text(row["summary"]),
                ", ".join(row["keywords"] or []), row["is_topic_relevant"], warning,
            ]
        )
        if warning:
            for cell in paragraphs_sheet[paragraphs_sheet.max_row]:
                cell.fill = WARNING_FILL
    _style_sheet(paragraphs_sheet, freeze="A2", wrap=True, widths={2: 42, 4: 28, 6: 80, 7: 80, 8: 60, 9: 36, 11: 36})

    keywords_sheet = workbook.create_sheet("대표키워드")
    keywords_sheet.append(["논문 ID", "제목", "정규화 키워드", "표시 키워드", "누적 빈도", "논문 점수"])
    for row in data["keywords"]:
        keywords_sheet.append(
            [
                row["paper_id"], paper_by_id[int(row["paper_id"])]["title"], row["keyword"],
                row["display_form"], row["frequency"], row["score"],
            ]
        )
    _style_sheet(keywords_sheet, freeze="A2", wrap=True, widths={2: 48, 3: 30, 4: 30})

    tables_sheet = workbook.create_sheet("표_추출")
    tables_sheet.append(["논문 ID", "제목", "표 ID", "표 제목", "표 OCR", "표 요약"])
    for row in data["tables"]:
        tables_sheet.append(
            [
                row["paper_id"], paper_by_id[int(row["paper_id"])]["title"], row["table_id"],
                row["table_title"], _excel_text(row["table_text"]), _excel_text(row["table_summary"]),
            ]
        )
    _style_sheet(tables_sheet, freeze="A2", wrap=True, widths={2: 48, 4: 36, 5: 80, 6: 60})
    return workbook


def _build_markdown(
    settings: Settings,
    documents: Sequence[ReviewDocument],
    data: dict[str, list[RowMapping]],
    output: Path,
) -> str:
    papers = {int(row["paper_id"]): row for row in data["papers"]}
    paragraphs = _group_by_paper(data["paragraphs"])
    keywords = _group_by_paper(data["keywords"])
    tables = _group_by_paper(data["tables"])
    lines = [
        "# 2편 OCR·레이아웃·키워드·요약 실측 결과",
        "",
        f"상세 원문은 `{output}`의 레이아웃_OCR·단락_요약 시트에서 확인한다.",
        "",
        "| 논문 | 페이지 | 블록 | 단락 | 표 | 대표 키워드 | 검수 화면 |",
        "| --- | ---: | ---: | ---: | ---: | --- | --- |",
    ]
    for document in documents:
        paper_id = int(document.paper_id or 0)
        paper = papers[paper_id]
        keyword_text = ", ".join(str(row["display_form"]) for row in keywords.get(paper_id, []))
        viewer = f"{settings.public_api_base_url.rstrip('/')}/documents/{document.document_id}/viewer"
        lines.append(
            f"| {paper['title']} | {len(document.pages)} | {len(document.blocks)} | "
            f"{len(paragraphs.get(paper_id, []))} | {len(tables.get(paper_id, []))} | "
            f"{keyword_text} | [열기]({viewer}) |"
        )
    lines.extend(["", "## LLM 요약 자동 경고", ""])
    total_warnings = 0
    total_paragraphs = 0
    for document in documents:
        paper_id = int(document.paper_id or 0)
        paper_rows = paragraphs.get(paper_id, [])
        warnings = [
            row for row in paper_rows if _summary_warning(str(row["summary"] or ""))
        ]
        total_warnings += len(warnings)
        total_paragraphs += len(paper_rows)
        orders = ", ".join(str(row["paragraph_order"]) for row in warnings) or "없음"
        lines.append(
            f"- **{papers[paper_id]['title']}**: {len(warnings)}/{len(paper_rows)}개 "
            f"경고(단락 순서: {orders})"
        )
    ratio = total_warnings / total_paragraphs * 100 if total_paragraphs else 0.0
    lines.append(
        f"- 전체 {total_warnings}/{total_paragraphs}개({ratio:.1f}%). 중국어·벵골어 혼입과 "
        "응답 형식 오염을 휴리스틱으로 표시했으며, 내용 정확성은 별도 사람 평가가 필요하다."
    )
    lines.extend(["", "## 관찰된 한계", ""])
    for document in documents:
        paper = papers[int(document.paper_id or 0)]
        lines.append(f"- **{paper['title']}**: {_quality_note(document, paper)}")
    lines.extend(
        [
            "",
            "- 노란색 행은 사람이 교정한 블록이다. 자동 인식 유형·원시 OCR과 현재 유형·교정 OCR을 나란히 남겼다.",
            "- 대표 키워드는 논문당 3~5개만 연결하며, 단락 키워드는 별도 열에 보존한다.",
            "- 이 결과는 학습 없이 PP-DocLayout-M, PP-OCRv5 mobile, Qwen2.5 7B, BGE-M3를 사용한 실측값이다.",
            "",
        ]
    )
    return "\n".join(lines)


def _quality_note(document: ReviewDocument, paper: RowMapping) -> str:
    notes: list[str] = []
    detected_types = [block.detected_block_type or block.block_type for block in document.blocks]
    if "title" not in detected_types:
        notes.append("자동 제목 영역 누락")
    if "abstract" not in detected_types or not str(paper["abstract"] or "").strip():
        notes.append("초록 영역/텍스트 누락")
    corrected = sum(block.review_status == "corrected" for block in document.blocks)
    if corrected:
        notes.append(f"{corrected}개 블록 검수 교정")
    if not notes:
        notes.append("치명적 메타데이터 누락 없음; OCR 철자·도형 내부 텍스트는 개별 확인 필요")
    return "; ".join(notes)


def _summary_warning(summary: str) -> str:
    warnings: list[str] = []
    if HAN_RE.search(summary):
        warnings.append("중국어/한자 혼입")
    if BENGALI_RE.search(summary):
        warnings.append("벵골어 문자 혼입")
    if FORMAT_POLLUTION_RE.search(summary):
        warnings.append("응답 형식 오염")
    return ", ".join(warnings)


def _group_by_paper(rows: Iterable[RowMapping]) -> dict[int, list[RowMapping]]:
    grouped: dict[int, list[RowMapping]] = {}
    for row in rows:
        grouped.setdefault(int(row["paper_id"]), []).append(row)
    return grouped


def _bbox_text(bbox: tuple[float, float, float, float] | None) -> str:
    return "" if bbox is None else ", ".join(f"{value:.2f}" for value in bbox)


def _excel_text(value: Any) -> str:
    if value is None:
        return ""
    rendered = str(value)
    if len(rendered) <= EXCEL_CELL_LIMIT:
        return rendered
    marker = "\n...[Excel 셀 길이 제한으로 잘림]"
    return rendered[: EXCEL_CELL_LIMIT - len(marker)] + marker


def _style_sheet(
    sheet: Worksheet,
    *,
    freeze: str,
    wrap: bool,
    widths: dict[int, float],
) -> None:
    sheet.freeze_panes = freeze
    sheet.auto_filter.ref = sheet.dimensions
    for cell in sheet[1]:
        cell.fill = HEADER_FILL
        cell.font = Font(color="FFFFFF", bold=True)
        cell.alignment = Alignment(horizontal="center", vertical="center")
    for row in sheet.iter_rows(min_row=2):
        for cell in row:
            cell.alignment = Alignment(vertical="top", wrap_text=wrap)
    for index in range(1, sheet.max_column + 1):
        sheet.column_dimensions[get_column_letter(index)].width = widths.get(index, 18)


if __name__ == "__main__":
    raise SystemExit(main())

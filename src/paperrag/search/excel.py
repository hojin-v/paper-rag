"""검색 결과(ResultBundle)를 xlsx 파일로 직렬화한다.

DESIGN.md §5.3에 정의된 시트 구성: ① 검색 결과 요약 ② 대표 논문 정보 ③ 대표 논문 섹션
④ 대표 논문 단락 ⑤ 연관 논문 정보 ⑥ 연관 논문 섹션 ⑦ 연관 논문 단락 ⑧ 표 데이터 ⑨ 표 셀.
(설계서 요약은 6개 범주로 묶어 표현하지만 실제 워크북은 섹션/단락/표 셀 시트가 별도로 나뉜다.)
기본값(모든 include_* 옵션 켜짐)에서는 9개 시트가 전부 만들어지지만, ResultBundle의
include_related=False/include_tables=False로 사용자가 산출물 구성을 좁히면 해당 시트
자체가 워크북에서 빠진다(⑤~⑦ 또는 ⑧~⑨). 대부분의 시트는 헤더 행 고정(freeze_panes)과
원문 셀 줄바꿈, 열 너비 자동 조정을 공통 적용한다 — 단, "표 셀" 시트는 표마다 열 구성이
달라 표 단위로 실제 행×열 그리드를 그대로 그리고 표 사이만 제목 행으로 구분한다
(전역 헤더 고정 대상이 아니다).
"""

from collections.abc import Iterable
from pathlib import Path
from typing import Any

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.table import Table, TableStyleInfo
from openpyxl.worksheet.worksheet import Worksheet

from paperrag.search.schemas import (
    PaperInfo,
    ParagraphInfo,
    ResultBundle,
    SectionInfo,
    TableInfo,
)

HEADER_FILL = PatternFill(fill_type="solid", fgColor="D9EAF7")
HEADER_FONT = Font(bold=True)


def build_excel(data: ResultBundle, out_path: str | Path) -> str:
    """ResultBundle 하나를 워크북 전체 시트로 조립해 out_path에 저장한다.

    SearchService.resolve()가 검색 결과를 확정한 직후 호출하며, 반환값(저장 경로
    문자열)은 그대로 search_results.excel_path로 저장돼 GET /result/{id}/excel
    다운로드 시 재사용된다.
    """
    path = Path(out_path)
    path.parent.mkdir(parents=True, exist_ok=True)

    workbook = Workbook()
    # 시트 1: 검색 결과 요약 — 질의/매칭/선정 사유를 한눈에 보는 시트. Workbook 생성 시
    # 기본으로 만들어지는 활성 시트를 그대로 재사용한다.
    summary_sheet = workbook.active
    summary_sheet.title = "검색 결과 요약"
    _write_summary(summary_sheet, data)

    # 시트 2: 대표 논문 정보 — 제목/저자/초록 등 메타데이터 전체.
    primary_sheet = workbook.create_sheet("대표 논문 정보")
    _write_paper_info(primary_sheet, data.primary_info)

    # 시트 3: 대표 논문 섹션 — 단락을 섹션 단위로 묶은 집계 뷰.
    primary_section_sheet = workbook.create_sheet("대표 논문 섹션")
    _write_sections(primary_section_sheet, data.primary_sections)

    # 시트 4: 대표 논문 단락 — 단락 단위 원문/정제문/요약/키워드.
    primary_paragraph_sheet = workbook.create_sheet("대표 논문 단락")
    _write_paragraphs(primary_paragraph_sheet, data.primary_paragraphs)

    paragraph_sheet_names = ["대표 논문 섹션", "대표 논문 단락"]

    # 시트 5~7: 연관 논문 정보/섹션/단락. data.include_related=False면(사용자가
    # 명시적으로 "연관 논문 제외"를 선택) 이 세 시트 자체를 만들지 않는다.
    # (연관 논문이 우연히 없는 경우와는 다르다 — 그때는 include_related=True인 채로
    # data.related_info=None만 되어 헤더만 있는 빈 시트가 그대로 유지된다.)
    if data.include_related:
        related_sheet = workbook.create_sheet("연관 논문 정보")
        _write_paper_info(
            related_sheet,
            data.related_info,
            relation_score=data.related_paper.score if data.related_paper else None,
            relation_reason=data.related_paper.reason if data.related_paper else None,
        )
        related_section_sheet = workbook.create_sheet("연관 논문 섹션")
        _write_sections(related_section_sheet, data.related_sections)
        related_paragraph_sheet = workbook.create_sheet("연관 논문 단락")
        _write_paragraphs(related_paragraph_sheet, data.related_paragraphs)
        paragraph_sheet_names.extend(["연관 논문 섹션", "연관 논문 단락"])

    # 시트 8~9: 표 데이터/표 셀. data.include_tables=False면 표 시트 자체를 만들지 않는다.
    tables_sheet: Worksheet | None = None
    if data.include_tables:
        tables_sheet = workbook.create_sheet("표 데이터")
        _write_tables(tables_sheet, data.tables)
        table_cells_sheet = workbook.create_sheet("표 셀")
        _write_table_cells(table_cells_sheet, data.tables)

    # 공통 서식: 대부분의 시트엔 헤더 강조 + 줄바꿈 + 열 너비 자동 조정을 적용한 뒤,
    # 원문/정제문이 긴 섹션·단락 시트와 표 내용 시트는 해당 열 너비를 고정폭으로 덮어써
    # 자동 조정 결과가 지나치게 좁아지는 것을 막는다. "표 셀" 시트는 표마다 열 구성이
    # 달라 전역 헤더 고정이 의미가 없으므로 별도의 가벼운 서식만 적용한다.
    for sheet in workbook.worksheets:
        if sheet.title == "표 셀":
            _style_table_cell_sheet(sheet)
        else:
            _style_sheet(sheet)
    for sheet_name in paragraph_sheet_names:
        _style_paragraph_sheet(workbook[sheet_name])
    if tables_sheet is not None:
        _style_table_sheet(tables_sheet)

    workbook.save(path)
    return str(path)


def _write_summary(sheet: Worksheet, data: ResultBundle) -> None:
    """"검색 결과 요약" 시트: 질의·매칭 방식·대표/연관 논문·선정 사유를 헤더+1행으로 요약한다."""
    _append_row(
        sheet,
        [
            "질의",
            "질의 추출 키워드",
            "매칭 키워드",
            "매칭 방식",
            "검색 설명",
            "대표 논문 제목",
            "연관 논문 제목",
            "대표 RAG 점수",
            "대표 선정 사유",
            "대표 관련도 설명",
            "연관 관계 점수",
            "연관 선정 사유",
            "연관 관련도 설명",
            "생성 일시",
        ],
    )
    related_title = data.related_paper.title if data.related_paper else ""
    _append_row(
        sheet,
        [
            data.query,
            ", ".join(data.query_keywords),
            data.matched_keyword,
            data.match_type,
            data.explanation,
            data.primary_paper.title,
            related_title,
            data.primary_paper.score,
            data.primary_paper.reason,
            data.primary_paper.relevance_summary,
            data.related_paper.score if data.related_paper else None,
            data.related_paper.reason if data.related_paper else None,
            data.related_paper.relevance_summary if data.related_paper else None,
            data.created_at.isoformat(timespec="seconds"),
        ],
    )


def _write_paper_info(
    sheet: Worksheet,
    paper: PaperInfo | None,
    *,
    relation_score: float | None = None,
    relation_reason: str | None = None,
) -> None:
    """"대표/연관 논문 정보" 시트 공용 작성 함수.

    relation_score/relation_reason이 주어지면(연관 논문 시트 호출 시) "연관 점수",
    "연관 사유" 열을 추가해 대표 논문 시트와 컬럼 구성을 다르게 한다. paper가
    None이면(연관 논문이 없는 경우) 헤더만 쓰고 데이터 행은 생략한다.
    """
    headers = [
        "논문 ID",
        "제목",
        "저자",
        "연도",
        "저널",
        "초록 원문",
        "초록 요약",
        "전문 링크",
        "키워드",
    ]
    include_relation = relation_score is not None or relation_reason is not None
    if include_relation:
        headers.extend(["연관 점수", "연관 사유"])
    _append_row(sheet, headers)
    if paper is None:
        return
    row: list[Any] = [
        paper.paper_id,
        paper.title,
        paper.authors,
        paper.published_year,
        paper.journal,
        paper.abstract,
        paper.abstract_summary,
        paper.full_text_link,
        ", ".join(paper.keywords),
    ]
    if include_relation:
        row.extend([relation_score, relation_reason])
    _append_row(sheet, row)


def _write_sections(sheet: Worksheet, sections: Iterable[SectionInfo]) -> None:
    """"대표/연관 논문 섹션" 시트: 단락을 섹션 단위로 합친 SectionInfo 목록을 행으로 나열한다."""
    _append_row(
        sheet,
        ["섹션 순서", "섹션명", "단락 수", "섹션 원문", "섹션 정제문", "섹션 요약", "키워드"],
    )
    for section in sections:
        _append_row(
            sheet,
            [
                section.section_order,
                section.section_name,
                section.paragraph_count,
                section.original_text,
                section.cleaned_text,
                section.summary,
                ", ".join(section.keywords),
            ],
        )


def _write_paragraphs(sheet: Worksheet, paragraphs: Iterable[ParagraphInfo]) -> None:
    """"대표/연관 논문 단락" 시트: is_topic_relevant=true인 단락을 순서대로 나열한다."""
    _append_row(sheet, ["단락 번호", "섹션명", "원문", "정제문", "요약", "키워드"])
    for paragraph in paragraphs:
        _append_row(
            sheet,
            [
                paragraph.paragraph_order,
                paragraph.section_name,
                paragraph.original_text,
                paragraph.cleaned_text,
                paragraph.summary,
                ", ".join(paragraph.keywords),
            ],
        )


def _write_tables(sheet: Worksheet, tables: Iterable[TableInfo]) -> None:
    """"표 데이터" 시트: 대표/연관 논문에 속한 표를 role로 구분해 표 단위(1행=표 1개)로 나열한다.

    "표 번호"는 "표 셀" 시트에서 같은 표의 실제 그리드를 찾아볼 수 있는 교차 참조용
    번호(그 시트에 표가 나열된 순서와 동일)다. "표 내용"은 검색·비교용 평면 텍스트로
    남겨두고, 사람이 바로 읽고 복사해 쓸 실제 행×열 그리드는 "표 셀" 시트에 그린다.
    """
    _append_row(sheet, ["구분", "표 번호", "표 제목", "표 내용", "표 요약"])
    for table_index, table in enumerate(tables, start=1):
        _append_row(
            sheet,
            [table.role, table_index, table.table_title, table.table_text, table.table_summary],
        )


def _write_table_cells(sheet: Worksheet, tables: Iterable[TableInfo]) -> None:
    """"표 셀" 시트: 표마다 실제 엑셀 "표"(Table) 객체로 등록해 바로 필터·정렬해 쓸 수 있게 한다.

    과거의 (구분/표번호/표제목/행/열/셀값) 평면 목록 대신, 표 사이를 볼드 제목 행
    ("[역할] 표 N — 제목")으로 구분하고 그 아래 실제 그리드를 `openpyxl.worksheet.table.Table`로
    등록한다 — 엑셀에서 열면 일반 셀 나열이 아니라 필터 화살표·줄무늬 서식이 있는
    진짜 표로 보인다. 표의 첫 데이터 행을 헤더 행으로 등록하는데(표 구조 HTML의
    <th>/<td> 구분은 이미 소실됐으므로 "첫 행 = 헤더"라는 관례적 추정), 엑셀 표는
    헤더 셀이 비어있거나 중복되면 열 때 복구 경고를 띄우므로 `_sanitize_table_headers`로
    미리 채우고 구분해 둔다.
    """
    for table_index, table in enumerate(tables, start=1):
        label = f"[{table.role}] 표 {table_index}"
        if table.table_title:
            label = f"{label} — {table.table_title}"
        _append_row(sheet, [label])
        _emphasize_row(sheet, sheet.max_row, 1, bold=True)

        rows = _parse_table_rows(table.table_text)
        if not rows:
            _append_row(sheet, [])
            continue

        header_row_num = sheet.max_row + 1
        max_cols = max(len(row) for row in rows)
        _append_row(sheet, _sanitize_table_headers(rows[0], max_cols))
        for row in rows[1:]:
            _append_row(sheet, row)
        last_row_num = sheet.max_row

        excel_table = Table(
            displayName=f"Tbl{table_index}",
            ref=f"A{header_row_num}:{get_column_letter(max_cols)}{last_row_num}",
        )
        excel_table.tableStyleInfo = TableStyleInfo(
            name="TableStyleMedium2", showRowStripes=True
        )
        sheet.add_table(excel_table)
        _append_row(sheet, [])


def _sanitize_table_headers(header_row: list[str], ncols: int) -> list[str]:
    """엑셀 표의 헤더 행 제약(빈 값·중복 이름 금지)을 만족하도록 헤더 셀을 채운다.

    ncols까지 부족한 칸은 빈 헤더로 채워야 하므로(엑셀 표 범위는 직사각형이어야
    함) "열N"으로 채우고, 실제 값이 있어도 비어 있거나 중복되면 마찬가지로 보정한다.
    """
    seen: dict[str, int] = {}
    sanitized: list[str] = []
    for index in range(1, ncols + 1):
        raw = header_row[index - 1].strip() if index <= len(header_row) else ""
        name = raw or f"열{index}"
        seen[name] = seen.get(name, 0) + 1
        if seen[name] > 1:
            name = f"{name}_{seen[name]}"
        sanitized.append(name)
    return sanitized


def _parse_table_rows(table_text: str) -> list[list[str]]:
    """수집 파이프라인 STEP 3에서 Markdown으로 직렬화해 저장한 table_text를 다시 행/열로 분해한다.

    "|"로 구분된 줄만 표 행으로 인식하고, 파이프가 없는 줄은 직전 셀에 이어붙이는
    줄바꿈 본문으로 취급한다(멀티라인 셀 대응). "TABLE "로 시작하는 줄을 만나면
    다음 표의 시작으로 보고 파싱을 중단한다(한 table_text에 여러 표 조각이 이어져
    있는 경우를 방어).
    """
    rows: list[list[str]] = []
    for raw_line in table_text.splitlines():
        line = raw_line.strip().strip("|").strip()
        if not line:
            continue
        if "|" in line:
            cells = [cell.strip() for cell in line.split("|")]
            if len(cells) >= 2:
                rows.append(cells)
            continue
        if line.upper().startswith("TABLE "):
            break
        if rows:
            rows[-1][0] = f"{rows[-1][0]} {line}".strip()
    return rows


def _append_row(sheet: Worksheet, values: list[Any]) -> None:
    """None 값을 빈 문자열로 바꿔 시트에 한 행을 추가하는 공용 헬퍼."""
    sheet.append(["" if value is None else value for value in values])


def _emphasize_row(
    sheet: Worksheet, row_index: int, ncols: int, *, bold: bool = False, fill: bool = False
) -> None:
    """지정한 행의 앞 ncols칸만 볼드/배경으로 강조한다("표 셀" 시트의 표 제목·헤더 행용)."""
    for column in range(1, ncols + 1):
        cell = sheet.cell(row=row_index, column=column)
        if bold:
            cell.font = HEADER_FONT
        if fill:
            cell.fill = HEADER_FILL


def _style_sheet(sheet: Worksheet) -> None:
    """모든 시트에 공통 적용하는 서식: 헤더 행 고정 + 헤더 강조 + 전체 셀 줄바꿈 + 열 너비 자동 조정.

    freeze_panes="A2"로 1행(헤더)을 고정해 스크롤해도 헤더가 보이게 하고,
    원문/정제문처럼 긴 텍스트가 셀 안에서 줄바꿈되어 표시되도록 wrap_text를 켠다.
    """
    sheet.freeze_panes = "A2"
    for cell in sheet[1]:
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = Alignment(wrap_text=True, vertical="top")
    for row in sheet.iter_rows(min_row=2):
        for cell in row:
            cell.alignment = Alignment(wrap_text=True, vertical="top")
    _auto_width(sheet)


def _style_table_cell_sheet(sheet: Worksheet) -> None:
    """"표 셀" 시트 전용 서식: 표마다 헤더가 달라 전역 헤더 고정(freeze_panes)이나

    1행 강조는 의미가 없으므로(표별 제목·헤더 행 강조는 `_write_table_cells`가 이미
    했다) 줄바꿈 + 열 너비 자동 조정만 적용한다.
    """
    for row in sheet.iter_rows():
        for cell in row:
            cell.alignment = Alignment(wrap_text=True, vertical="top")
    _auto_width(sheet)


def _style_paragraph_sheet(sheet: Worksheet) -> None:
    """섹션/단락 시트의 원문(C열)·정제문(D열)은 자동 너비 대신 60폭으로 고정해 가독성을 확보한다."""
    for column in ("C", "D"):
        sheet.column_dimensions[column].width = 60


def _style_table_sheet(sheet: Worksheet) -> None:
    """표 데이터 시트의 표 내용(D열, "표 번호" 열 추가로 한 칸 밀림)도 같은 이유로 60폭으로 고정한다."""
    sheet.column_dimensions["D"].width = 60


def _auto_width(sheet: Worksheet) -> None:
    """각 열에서 가장 긴 줄(멀티라인 셀 고려) 길이에 맞춰 열 너비를 추정하되 40을 상한으로 둔다.

    이후 _style_paragraph_sheet/_style_table_sheet가 특정 긴 텍스트 열의 너비를
    다시 고정값으로 덮어써, 자동 조정이 과도하게 좁히거나 넓히는 것을 보정한다.
    """
    for column_index, column_cells in enumerate(sheet.columns, start=1):
        max_length = 0
        for cell in column_cells:
            value = str(cell.value or "")
            longest_line = max((len(line) for line in value.splitlines()), default=0)
            max_length = max(max_length, longest_line)
        sheet.column_dimensions[get_column_letter(column_index)].width = min(max_length + 2, 40)

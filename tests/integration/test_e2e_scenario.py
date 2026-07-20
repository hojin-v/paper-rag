import asyncio
import hashlib
import importlib.util
import math
from collections.abc import Iterator, Sequence
from dataclasses import dataclass
from io import BytesIO
from pathlib import Path
from typing import Any

import httpx
import pytest
from openpyxl import load_workbook
from sqlalchemy import create_engine, text
from sqlalchemy.engine import Engine

pgserver = pytest.importorskip("pgserver")
pymupdf = pytest.importorskip("pymupdf")

from paperrag.config import Settings  # noqa: E402
from paperrag.ingest.embeddings import EmbeddingClient  # noqa: E402
from paperrag.ingest.layout.simple_backend import SimplePyMuPDFBackend  # noqa: E402
from paperrag.ingest.llm_enrich import LLMClient  # noqa: E402
from paperrag.ingest.pipeline import IngestPipeline  # noqa: E402
from paperrag.ingest.repository import PostgresIngestRepository  # noqa: E402
from paperrag.search.api import app, get_service  # noqa: E402
from paperrag.search.repository import PostgresSearchRepository  # noqa: E402
from paperrag.search.schemas import SearchMatched, SearchSuggest  # noqa: E402
from paperrag.search.service import SearchService  # noqa: E402


if importlib.util.find_spec("httpx2") is None:

    class TestClient:
        __test__ = False

        def __init__(self, asgi_app: Any) -> None:
            self.asgi_app = asgi_app

        def __enter__(self) -> "TestClient":
            return self

        def __exit__(self, exc_type: object, exc: object, traceback: object) -> None:
            return None

        def get(self, url: str, **kwargs: Any) -> httpx.Response:
            return asyncio.run(self._request("GET", url, **kwargs))

        def post(self, url: str, **kwargs: Any) -> httpx.Response:
            return asyncio.run(self._request("POST", url, **kwargs))

        async def _request(self, method: str, url: str, **kwargs: Any) -> httpx.Response:
            transport = httpx.ASGITransport(app=self.asgi_app)
            async with httpx.AsyncClient(
                transport=transport,
                base_url="http://testserver",
            ) as client:
                response = await client.request(method, url, **kwargs)
                await response.aread()
                return response

else:
    from fastapi.testclient import TestClient


@dataclass(frozen=True)
class PaperSpec:
    key: str
    title: str
    authors: str
    year: int
    section: str
    body_lines: tuple[str, ...]
    references: tuple[str, ...]


@dataclass(frozen=True)
class E2EContext:
    engine: Engine
    settings: Settings
    llm: "ScriptedLLM"
    embedder: "ControlledEmbedding"
    paper_ids: dict[str, int]


class ScriptedLLM(LLMClient):
    def generate_json(self, prompt: str, schema_hint: str) -> dict[str, Any]:
        if "사용자의 자연어 질의" in prompt:
            if "이상탐지" in prompt:
                return {"keywords": ["스마트팩토리", "이상탐지"]}
            if "예지보전" in prompt:
                return {"keywords": ["예지보전"]}
            return {"keywords": ["RAG"]}

        keywords = _paper_keywords_for(prompt)
        if "입력 단락:" in prompt:
            original = prompt.rsplit("입력 단락:", 1)[1].strip()
            return {
                "cleaned_text": original,
                "summary": "요약: " + original[:80],
                "keywords": keywords[:2],
                "is_topic_relevant": True,
            }
        if "대표 키워드" in prompt:
            return {"keywords": keywords}
        if "논문 표 내용" in prompt:
            return {"summary": "테스트 표 요약"}
        return {"keywords": keywords}


class ControlledEmbedding(EmbeddingClient):
    dim = 1024

    def __init__(self) -> None:
        forecast = _unit([0.95, math.sqrt(1.0 - 0.95**2)])
        diagnosis = _unit([0.8, 0.6])
        anomaly = _unit([0.72, math.sqrt(1.0 - 0.72**2)])
        smart_factory = _unit([0.65, 0.45, 0.61])
        deep_learning = _unit([0.55, 0.2, 0.81])
        rag = _unit([0.1, 0.0, math.sqrt(1.0 - 0.1**2)])

        self.known: dict[str, list[float]] = {
            "예지보전": _unit([1.0]),
            "예측 유지보수": forecast,
            "predictive maintenance": forecast,
            "설비 진단": diagnosis,
            "equipment diagnosis": diagnosis,
            "이상탐지": anomaly,
            "anomaly detection": anomaly,
            "스마트팩토리": smart_factory,
            "smart factory": smart_factory,
            "딥러닝": deep_learning,
            "deep learning": deep_learning,
            "rag": rag,
            "retrieval augmented": rag,
            "검색 증강 생성": rag,
            "llm": rag,
        }

    def embed(self, texts: list[str]) -> list[list[float]]:
        return [self._embed_one(text) for text in texts]

    def _embed_one(self, text_value: str) -> list[float]:
        lowered = text_value.lower()
        vectors: list[list[float]] = []
        for term, vector in self.known.items():
            count = lowered.count(term.lower())
            vectors.extend(vector for _ in range(count))
        if not vectors:
            return _hash_vector(text_value, self.dim)
        averaged = [
            sum(vector[index] for vector in vectors) / len(vectors)
            for index in range(self.dim)
        ]
        return _unit(averaged, self.dim)


PAPER_SPECS = (
    PaperSpec(
        key="paper1",
        title="Deep Learning based Anomaly Detection in Smart Factory",
        authors="Kim H.; Lee S.",
        year=2024,
        section="Introduction",
        body_lines=(
            "This paper studies anomaly detection for smart factory production lines.",
            "Deep learning models identify subtle anomaly detection signals in sensors.",
            "The anomaly detection method supports predictive maintenance planning.",
            "Smart factory monitoring combines vision logs and time-series measurements.",
            "Deep learning anomaly detection shows robust manufacturing defect results.",
            "The smart factory system recommends maintenance windows before failure.",
        ),
        references=(
            "[1] Reference material about factory monitoring.",
            "[2] Reference material about anomaly detection.",
        ),
    ),
    PaperSpec(
        key="paper2",
        title="Predictive Maintenance for Manufacturing Equipment",
        authors="Park J.; Choi M.",
        year=2023,
        section="Methods",
        body_lines=(
            "Predictive maintenance models estimate remaining useful life for equipment.",
            "Equipment diagnosis uses vibration, temperature, and operation logs.",
            "The predictive maintenance workflow prioritizes repair tickets.",
            "Predictive maintenance reduces manufacturing equipment failures.",
            "Equipment diagnosis groups similar fault signatures for engineers.",
            "A predictive maintenance dashboard summarizes equipment diagnosis outcomes.",
        ),
        references=(
            "[1] Reference material about maintenance scheduling.",
            "[2] Reference material about equipment diagnosis.",
        ),
    ),
    PaperSpec(
        key="paper3",
        title="Retrieval Augmented Generation Survey",
        authors="Jeong A.; Han B.",
        year=2024,
        section="Survey",
        body_lines=(
            "Retrieval augmented generation connects a retriever with an LLM.",
            "RAG systems ground answers in external document collections.",
            "The retrieval augmented survey compares chunking and reranking strategies.",
            "LLM applications benefit from citation-aware RAG retrieval.",
            "RAG evaluation covers answer faithfulness and source coverage.",
            "Future work studies multimodal retrieval augmented pipelines.",
        ),
        references=(
            "[1] Reference material about RAG.",
            "[2] Reference material about language models.",
        ),
    ),
)


@pytest.fixture(scope="session")
def e2e_context(pg_dsn: str, tmp_path_factory: pytest.TempPathFactory) -> Iterator[E2EContext]:
    sqlalchemy_dsn = _sqlalchemy_dsn(pg_dsn)
    engine = create_engine(sqlalchemy_dsn, pool_pre_ping=True)
    llm = ScriptedLLM()
    embedder = ControlledEmbedding()
    settings = Settings(
        _env_file=None,
        database_url=sqlalchemy_dsn,
        result_dir=tmp_path_factory.mktemp("results"),
        paragraph_min_chars=40,
        paragraph_max_chars=1500,
        search_suggestion_limit=3,
        search_similarity_threshold=0.6,
        relation_top_k=20,
    )
    ingest_repo = PostgresIngestRepository(settings, engine)
    pipeline = IngestPipeline(
        ingest_repo,
        SimplePyMuPDFBackend(),
        llm,
        embedder,
        settings=settings,
    )
    pdf_dir = tmp_path_factory.mktemp("pdfs")
    paper_ids: dict[str, int] = {}

    for spec in PAPER_SPECS:
        pdf_path = _write_pdf(pdf_dir, spec)
        report = pipeline.run(str(pdf_path))
        assert report.paper_id is not None
        assert not report.errors
        assert report.totals["paragraphs"] >= 1
        assert report.totals["keywords"] >= 3
        paper_ids[spec.key] = report.paper_id

    _assert_seed_relation(engine, paper_ids["paper1"], paper_ids["paper2"])

    try:
        yield E2EContext(
            engine=engine,
            settings=settings,
            llm=llm,
            embedder=embedder,
            paper_ids=paper_ids,
        )
    finally:
        engine.dispose()


@pytest.fixture
def search_service(e2e_context: E2EContext) -> SearchService:
    return SearchService(
        PostgresSearchRepository(e2e_context.settings, e2e_context.engine),
        e2e_context.llm,
        e2e_context.embedder,
        e2e_context.settings,
    )


@pytest.fixture
def api_client(search_service: SearchService) -> Iterator[TestClient]:
    async def override_service() -> SearchService:
        return search_service

    app.dependency_overrides[get_service] = override_service
    with TestClient(app) as client:
        yield client
    app.dependency_overrides.clear()


def test_exact_match_search_and_excel(
    e2e_context: E2EContext,
    search_service: SearchService,
) -> None:
    # use_llm=True: 이 테스트는 ScriptedLLM의 프롬프트별 키워드 시나리오(자연어 이해 경로)를
    # 검증하는 것이 목적이므로, 기본값인 형태소 분석 빠른 경로가 아니라 명시적으로 AI 경로를 켠다.
    result = search_service.search(
        "스마트팩토리에서 이상탐지를 위한 딥러닝 기반 예측 유지보수 논문을 찾고 싶다",
        use_llm=True,
    )

    assert isinstance(result, SearchMatched)
    assert result.match_type == "exact"
    assert result.primary_paper.paper_id == e2e_context.paper_ids["paper1"]
    assert result.related_paper is not None
    assert result.related_paper.paper_id == e2e_context.paper_ids["paper2"]

    excel_path = search_service.result_excel_path(result.result_id)
    assert excel_path is not None
    _assert_excel(Path(excel_path), result.primary_paper.title)


def test_similar_keyword_suggest_select_and_excel(
    e2e_context: E2EContext,
    search_service: SearchService,
) -> None:
    # use_llm=True: ScriptedLLM의 "예지보전" 시나리오(자연어 이해 경로)를 검증한다.
    suggestion = search_service.search("예지보전 관련 논문", use_llm=True)

    assert isinstance(suggestion, SearchSuggest)
    candidates = {candidate.keyword: candidate for candidate in suggestion.candidates}
    assert "예측 유지보수" in candidates
    assert candidates["예측 유지보수"].similarity >= 0.6
    assert "RAG" not in candidates

    selected = search_service.select(
        suggestion.session_id,
        candidates["예측 유지보수"].keyword_id,
    )

    assert selected.match_type == "selected"
    assert selected.primary_paper.paper_id == e2e_context.paper_ids["paper2"]
    assert selected.related_paper is not None
    assert selected.related_paper.paper_id == e2e_context.paper_ids["paper1"]

    excel_path = search_service.result_excel_path(selected.result_id)
    assert excel_path is not None
    _assert_excel(Path(excel_path), selected.primary_paper.title)


def test_section_query_filters_paragraph_sheet_against_real_postgres(
    e2e_context: E2EContext,
    search_service: SearchService,
) -> None:
    """section_query가 실제 Postgres에서 ILIKE로 올바르게 단락을 좁히는지 확인한다.

    (InMemorySearchRepository만으로는 잡을 수 없었던 PostgresSearchRepository의
    AmbiguousParameter 타입 캐스팅 버그를 이 테스트가 실제로 발견했다 — 회귀 방지용.)
    """
    query = "스마트팩토리에서 이상탐지를 위한 딥러닝 기반 예측 유지보수 논문을 찾고 싶다"

    unfiltered = search_service.search(query, use_llm=True)
    assert isinstance(unfiltered, SearchMatched)
    unfiltered_rows = _paragraph_row_count(search_service, unfiltered.result_id)
    assert unfiltered_rows > 1  # 헤더 + 최소 1개 데이터 행

    # SimplePyMuPDFBackend는 section_header 블록 타입을 만들지 않으므로, 이 픽스처의
    # 모든 논문 단락은 paragraphs.build_paragraphs의 기본 섹션명("본문") 그대로 저장된다.
    # 부분 일치 필터를 걸어도 전량이 그대로 남아야 한다.
    matching = search_service.search(query, use_llm=True, section_query="본문")
    assert isinstance(matching, SearchMatched)
    assert _paragraph_row_count(search_service, matching.result_id) == unfiltered_rows

    # 존재하지 않는 섹션으로 필터하면 헤더 행만 남아야 한다.
    empty = search_service.search(query, use_llm=True, section_query="이런섹션은없다")
    assert isinstance(empty, SearchMatched)
    assert _paragraph_row_count(search_service, empty.result_id) == 1


def test_references_are_excluded_from_paragraphs(e2e_context: E2EContext) -> None:
    with e2e_context.engine.begin() as connection:
        rows = connection.execute(
            text(
                """
                SELECT original_text
                FROM paragraphs
                WHERE paper_id = :paper_id
                ORDER BY paragraph_order
                """
            ),
            {"paper_id": e2e_context.paper_ids["paper1"]},
        ).scalars().all()

    paragraph_text = "\n".join(str(row) for row in rows)
    assert "References" not in paragraph_text
    assert "Reference material" not in paragraph_text


def test_api_level_search_select_and_excel(api_client: TestClient) -> None:
    # use_llm=True: API 계층에서도 ScriptedLLM 자연어 이해 경로가 그대로 동작하는지 확인한다.
    matched_response = api_client.post(
        "/search",
        json={
            "query": "스마트팩토리에서 이상탐지를 위한 딥러닝 기반 예측 유지보수 논문을 찾고 싶다",
            "use_llm": True,
        },
    )

    assert matched_response.status_code == 200
    matched = matched_response.json()
    assert matched["status"] == "matched"
    assert matched["match_type"] == "exact"

    excel_response = api_client.get(f"/result/{matched['result_id']}/excel")
    assert excel_response.status_code == 200
    assert excel_response.headers["content-type"].startswith(
        "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )
    assert load_workbook(BytesIO(excel_response.content)).sheetnames == _expected_sheet_names()

    suggest_response = api_client.post(
        "/search", json={"query": "예지보전 관련 논문", "use_llm": True}
    )
    assert suggest_response.status_code == 200
    suggestion = suggest_response.json()
    assert suggestion["status"] == "suggest"
    selected_keyword_id = next(
        candidate["keyword_id"]
        for candidate in suggestion["candidates"]
        if candidate["keyword"] == "예측 유지보수"
    )

    select_response = api_client.post(
        "/search/select",
        json={
            "session_id": suggestion["session_id"],
            "keyword_id": selected_keyword_id,
        },
    )

    assert select_response.status_code == 200
    selected = select_response.json()
    assert selected["status"] == "matched"
    assert selected["match_type"] == "selected"

    selected_excel_response = api_client.get(f"/result/{selected['result_id']}/excel")
    assert selected_excel_response.status_code == 200
    assert selected_excel_response.headers["content-type"].startswith(
        "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )
    assert load_workbook(BytesIO(selected_excel_response.content)).sheetnames == (
        _expected_sheet_names()
    )


def _paragraph_row_count(search_service: SearchService, result_id: str) -> int:
    excel_path = search_service.result_excel_path(result_id)
    assert excel_path is not None
    return load_workbook(excel_path)["대표 논문 단락"].max_row


def _paper_keywords_for(text_value: str) -> list[str]:
    lowered = text_value.lower()
    if "retrieval augmented" in lowered:
        return ["RAG", "검색 증강 생성", "LLM"]
    if "anomaly detection" in lowered or "smart factory" in lowered:
        return ["스마트팩토리", "이상탐지", "딥러닝", "예측 유지보수"]
    if "predictive maintenance" in lowered:
        return ["예측 유지보수", "이상탐지", "설비 진단"]
    return []


def _unit(values: Sequence[float], dim: int = 1024) -> list[float]:
    padded = [0.0] * dim
    for index, value in enumerate(values[:dim]):
        padded[index] = float(value)
    norm = math.sqrt(sum(value * value for value in padded))
    if norm == 0.0:
        return padded
    return [value / norm for value in padded]


def _hash_vector(text_value: str, dim: int) -> list[float]:
    values: list[float] = []
    for index in range(8):
        digest = hashlib.sha256(f"{text_value}\0{index}".encode("utf-8")).digest()
        integer = int.from_bytes(digest[:4], "big", signed=False)
        values.append((integer / 2**32) * 2.0 - 1.0)
    return _unit(values, dim)


def _sqlalchemy_dsn(dsn: str) -> str:
    if dsn.startswith("postgresql+psycopg://"):
        return dsn
    return dsn.replace("postgresql://", "postgresql+psycopg://", 1)


def _write_pdf(pdf_dir: Path, spec: PaperSpec) -> Path:
    path = pdf_dir / f"{spec.key}.pdf"
    document = pymupdf.open()
    page = document.new_page(width=595, height=842)
    y = 72
    lines = [
        (spec.title, 16),
        (spec.authors, 11),
        (str(spec.year), 11),
        (spec.section, 12),
        *[(line, 11) for line in spec.body_lines],
        ("References", 12),
        *[(line, 11) for line in spec.references],
    ]
    for line, font_size in lines:
        page.insert_text((72, y), line, fontsize=font_size)
        y += 24
    document.save(path)
    document.close()
    return path


def _assert_seed_relation(engine: Engine, paper1_id: int, paper2_id: int) -> None:
    with engine.begin() as connection:
        row = connection.execute(
            text(
                """
                SELECT source_paper_id, related_paper_id, relation_score
                FROM paper_relations
                WHERE (source_paper_id = :paper1_id AND related_paper_id = :paper2_id)
                   OR (source_paper_id = :paper2_id AND related_paper_id = :paper1_id)
                ORDER BY relation_score DESC
                LIMIT 1
                """
            ),
            {"paper1_id": paper1_id, "paper2_id": paper2_id},
        ).mappings().first()

    assert row is not None
    assert float(row["relation_score"]) > 0.6


def _assert_excel(path: Path, expected_primary_title: str) -> None:
    assert path.exists()
    workbook = load_workbook(path)
    assert workbook.sheetnames == _expected_sheet_names()

    summary = workbook["검색 결과 요약"]
    assert summary["A2"].value
    assert summary["B2"].value
    assert summary["D2"].value in {"exact", "similar", "selected"}
    assert summary["E2"].value
    assert summary["F2"].value == expected_primary_title
    assert summary["I2"].value

    primary_sections = workbook["대표 논문 섹션"]
    assert primary_sections["B2"].value
    assert primary_sections["D2"].value

    primary_paragraphs = workbook["대표 논문 단락"]
    assert primary_paragraphs["C2"].value
    assert primary_paragraphs["E2"].value


def _expected_sheet_names() -> list[str]:
    return [
        "검색 결과 요약",
        "대표 논문 정보",
        "대표 논문 섹션",
        "대표 논문 단락",
        "연관 논문 정보",
        "연관 논문 섹션",
        "연관 논문 단락",
        "표 데이터",
        "표 셀",
    ]

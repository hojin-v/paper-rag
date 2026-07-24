from datetime import UTC, datetime
from pathlib import Path

from openpyxl import load_workbook

from paperrag.search.excel import build_excel
from paperrag.search.schemas import (
    PaperInfo,
    PaperSummary,
    ParagraphInfo,
    ResultBundle,
    SectionInfo,
    TableInfo,
)


def test_build_excel_writes_sections_paragraphs_and_table_cells(tmp_path: Path) -> None:
    out_path = tmp_path / "result.xlsx"

    excel_path = build_excel(_bundle(), out_path)

    workbook = load_workbook(excel_path)
    assert workbook.sheetnames == [
        "검색 결과 요약",
        "대표 논문 정보",
        "대표 논문 섹션",
        "대표 논문 단락",
        "연관 논문 정보",
        "연관 논문 섹션",
        "연관 논문 단락",
        "표 데이터",
        "표 셀",
    ]

    summary = workbook["검색 결과 요약"]
    assert summary.freeze_panes == "A2"
    assert summary["A1"].font.bold
    assert summary["A1"].value == "질의"
    assert summary["A2"].value == "RAG 관련 논문"
    assert summary["B2"].value == "RAG, 검색"
    assert summary["C2"].value == "RAG"
    assert summary["F2"].value == "Primary RAG Paper"
    assert summary["G2"].value == "Related OCR Paper"
    assert summary["H1"].value == "대표 RAG 점수"
    assert summary["H2"].value == 0.89
    assert summary["J1"].value == "대표 관련도 설명"
    assert summary["J2"].value == "대표 논문 관련도 설명"
    assert summary["K1"].value == "연관 관계 점수"
    assert summary["K2"].value == 0.77
    assert summary["L2"].value == "겹치는 키워드: RAG"
    assert summary["M1"].value == "연관 관련도 설명"
    assert summary["M2"].value == "연관 논문 관련도 설명"

    primary_info = workbook["대표 논문 정보"]
    assert primary_info["A1"].value == "논문 ID"
    assert primary_info["B2"].value == "Primary RAG Paper"
    assert primary_info["F2"].value == "대표 초록 원문"
    assert primary_info["I2"].value == "RAG, 검색"

    related_info = workbook["연관 논문 정보"]
    assert related_info["J1"].value == "연관 점수"
    assert related_info["J2"].value == 0.77
    assert related_info["K2"].value == "겹치는 키워드: RAG"

    primary_sections = workbook["대표 논문 섹션"]
    assert primary_sections["B2"].value == "Introduction"
    assert primary_sections["D2"].value == "원문 단락"
    assert primary_sections["F2"].value == "요약"

    primary_paragraphs = workbook["대표 논문 단락"]
    assert primary_paragraphs["A1"].value == "단락 번호"
    assert primary_paragraphs["C2"].value == "원문 단락"
    assert primary_paragraphs["D2"].value == "정제 단락"
    assert primary_paragraphs.column_dimensions["C"].width == 60
    assert primary_paragraphs.column_dimensions["D"].width == 60

    tables = workbook["표 데이터"]
    assert tables["A1"].value == "구분"
    assert tables["B1"].value == "표 번호"
    assert tables["A2"].value == "대표"
    assert tables["B2"].value == 1
    assert tables["D2"].value == "metric | value\nf1 | 0.90"

    table_cells = workbook["표 셀"]
    assert table_cells["A1"].value == "[대표] 표 1 — Table 1"
    assert table_cells["A1"].font.bold
    assert table_cells["A2"].value == "metric"
    assert table_cells["B2"].value == "value"
    assert table_cells["A3"].value == "f1"
    assert table_cells["B3"].value == "0.90"
    assert "Tbl1" in table_cells.tables
    assert table_cells.tables["Tbl1"].ref == "A2:B3"
    assert table_cells["B3"].value == "0.90"
    assert not table_cells["A3"].font.bold


def test_table_cells_sheet_sanitizes_ragged_and_duplicate_headers(tmp_path: Path) -> None:
    """빈 헤더·중복 헤더·행마다 다른 열 수를 가진 표도 유효한 엑셀 표 범위로 등록돼야 한다."""
    out_path = tmp_path / "result.xlsx"
    bundle = _bundle().model_copy(
        update={
            "tables": [
                TableInfo(
                    role="대표",
                    table_title="Ragged Table",
                    # 헤더 행: 빈 칸 하나, 데이터 행은 열 수가 제각각(엑셀 표는 직사각형이어야
                    # 하므로 최대 열 수 3에 맞춰 범위가 정해져야 한다).
                    table_text="a |  | 값\nx | y | z\np | q",
                    table_summary="표 요약",
                )
            ]
        }
    )

    excel_path = build_excel(bundle, out_path)

    workbook = load_workbook(excel_path)
    sheet = workbook["표 셀"]
    assert sheet["A2"].value == "a"
    assert sheet["B2"].value == "열2"
    assert sheet["C2"].value == "값"
    assert "Tbl1" in sheet.tables
    assert sheet.tables["Tbl1"].ref == "A2:C4"


def test_build_excel_excludes_related_sheets_when_disabled(tmp_path: Path) -> None:
    """include_related=False면 연관 논문 관련 시트 3개가 워크북에서 아예 빠져야 한다."""
    out_path = tmp_path / "result.xlsx"
    bundle = _bundle().model_copy(update={"include_related": False})

    excel_path = build_excel(bundle, out_path)

    workbook = load_workbook(excel_path)
    assert workbook.sheetnames == [
        "검색 결과 요약",
        "대표 논문 정보",
        "대표 논문 섹션",
        "대표 논문 단락",
        "표 데이터",
        "표 셀",
    ]


def test_build_excel_excludes_table_sheets_when_disabled(tmp_path: Path) -> None:
    """include_tables=False면 표 관련 시트 2개가 워크북에서 아예 빠져야 한다."""
    out_path = tmp_path / "result.xlsx"
    bundle = _bundle().model_copy(update={"include_tables": False})

    excel_path = build_excel(bundle, out_path)

    workbook = load_workbook(excel_path)
    assert workbook.sheetnames == [
        "검색 결과 요약",
        "대표 논문 정보",
        "대표 논문 섹션",
        "대표 논문 단락",
        "연관 논문 정보",
        "연관 논문 섹션",
        "연관 논문 단락",
    ]


def test_build_excel_excludes_related_and_table_sheets_when_both_disabled(
    tmp_path: Path,
) -> None:
    """둘 다 끄면 대표 논문 관련 시트 4개만 남아야 한다."""
    out_path = tmp_path / "result.xlsx"
    bundle = _bundle().model_copy(update={"include_related": False, "include_tables": False})

    excel_path = build_excel(bundle, out_path)

    workbook = load_workbook(excel_path)
    assert workbook.sheetnames == [
        "검색 결과 요약",
        "대표 논문 정보",
        "대표 논문 섹션",
        "대표 논문 단락",
    ]


def _bundle() -> ResultBundle:
    return ResultBundle(
        result_id="r-20260704-abcdef12",
        query="RAG 관련 논문",
        query_keywords=["RAG", "검색"],
        matched_keyword="RAG",
        match_type="exact",
        explanation="RAG 키워드로 대표 논문을 선택했습니다.",
        primary_paper=PaperSummary(
            paper_id=10,
            title="Primary RAG Paper",
            authors="Kim; Lee",
            published_year=2025,
            journal="Journal",
            abstract="대표 초록 원문",
            full_text_link="https://example.test/primary",
            keywords=["RAG", "검색"],
            score=0.89,
            reason="대표 점수=0.890",
            relevance_summary="대표 논문 관련도 설명",
        ),
        related_paper=PaperSummary(
            paper_id=30,
            title="Related OCR Paper",
            authors="Choi",
            published_year=2026,
            journal="Related",
            abstract="연관 초록 원문",
            full_text_link="https://example.test/related",
            keywords=["OCR"],
            score=0.77,
            reason="겹치는 키워드: RAG",
            relevance_summary="연관 논문 관련도 설명",
        ),
        primary_info=PaperInfo(
            paper_id=10,
            title="Primary RAG Paper",
            authors="Kim; Lee",
            published_year=2025,
            journal="Journal",
            abstract="대표 초록 원문",
            abstract_summary="대표 초록 요약",
            full_text_link="https://example.test/primary",
            keywords=["RAG", "검색"],
        ),
        related_info=PaperInfo(
            paper_id=30,
            title="Related OCR Paper",
            authors="Choi",
            published_year=2026,
            journal="Related",
            abstract="연관 초록 원문",
            abstract_summary="연관 초록 요약",
            full_text_link="https://example.test/related",
            keywords=["OCR"],
        ),
        primary_paragraphs=[
            ParagraphInfo(
                paragraph_order=1,
                section_name="Introduction",
                original_text="원문 단락",
                cleaned_text="정제 단락",
                summary="요약",
                keywords=["RAG"],
            )
        ],
        related_paragraphs=[
            ParagraphInfo(
                paragraph_order=1,
                section_name="Related",
                original_text="연관 원문",
                cleaned_text="연관 정제",
                summary="연관 요약",
                keywords=["OCR"],
            )
        ],
        primary_sections=[
            SectionInfo(
                section_order=1,
                section_name="Introduction",
                paragraph_count=1,
                original_text="원문 단락",
                cleaned_text="정제 단락",
                summary="요약",
                keywords=["RAG"],
            )
        ],
        related_sections=[
            SectionInfo(
                section_order=1,
                section_name="Related",
                paragraph_count=1,
                original_text="연관 원문",
                cleaned_text="연관 정제",
                summary="연관 요약",
                keywords=["OCR"],
            )
        ],
        tables=[
            TableInfo(
                role="대표",
                table_title="Table 1",
                table_text="metric | value\nf1 | 0.90",
                table_summary="표 요약",
            )
        ],
        created_at=datetime(2026, 7, 4, 2, 30, tzinfo=UTC),
    )
